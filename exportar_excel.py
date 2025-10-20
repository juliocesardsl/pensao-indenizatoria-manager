import sqlite3
import openpyxl
import tkinter as tk
import os
import sys
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, Alignment, Border, Side

# ================================================================================================= #
# FUNÇÃO AUXILIAR PARA CAMINHOS DE ARQUIVOS (PARA O EXECUTÁVEL)
# ================================================================================================= #
def resource_path(relative_path):
    """ Obtém o caminho absoluto para o recurso, funciona para dev e para PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def _get_save_path(default_filename, parent=None):
    """Abre uma janela para o usuário escolher onde salvar o arquivo Excel.

    parent: janela Tk/Toplevel opcional para manter o diálogo em frente.
    """
    # Se nenhum parent for fornecido, tente usar a raiz padrão do tkinter
    if parent is None:
        try:
            parent = tk._default_root
        except Exception:
            parent = None

    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        initialfile=default_filename,
        title="Salvar Relatório Excel"
        , parent=parent
    )
    return filepath

def _apply_header_style(sheet):
    """Aplica um estilo padrão aos cabeçalhos da planilha."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="a1970c", end_color="a1970c", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def exportar_beneficiarios_excel():
    """Exporta a lista completa de beneficiários para um arquivo Excel."""
    save_path = _get_save_path("Relatorio_Beneficiarios.xlsx")
    if not save_path:
        return

    try:
        conn = sqlite3.connect(resource_path('banco.db'))
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                b.id, b.nome_completo, b.menor_ou_incapaz, b.cpf, b.identidade, b.orgao_emissor, b.email, r.nome_completo, b.endereco, b.cep, b.telefone, 
                b.numero_processo_judicial, b.numero_processo_sei, b.origem_decisao, b.numero_vara,
                b.agencia, b.numero_conta, b.digitoconta, b.numero_banco, b.descricao_banco, b.tipo_conta,
                b.data_decisao, b.data_oficioPGDF, b.observacoes
            FROM beneficiarios b
            LEFT JOIN representantes_legais r ON b.cpf = r.cpf_beneficiario
            ORDER BY b.nome_completo
        """)
        beneficiarios = cursor.fetchall()
        conn.close()

        if not beneficiarios:
            messagebox.showinfo("Informação", "Nenhum beneficiário cadastrado para exportar.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Beneficiários"

        headers = [
            "ID", "Nome", "Menor/Incapaz", "CPF", "Identidade", "Órgão Emissor", "Email", "Responsável Legal", "Endereço", "CEP", "Telefone",
            "Proc. Judicial", "Proc. SEI", "Origem Decisão", "Nº Vara",
            "Agência", "Conta", "Dígito", "Nº Banco", "Descrição Banco", "Tipo Conta",
            "Data Decisão", "Data PGDF", "Ementa da Decisão"
        ]
        ws.append(headers)
        _apply_header_style(ws)

        for beneficiario in beneficiarios:
            beneficiario_lista = list(beneficiario)
            beneficiario_lista[2] = "Sim" if beneficiario_lista[2] else "Não"
            ws.append(beneficiario_lista)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(save_path)
        messagebox.showinfo("Sucesso", f"Lista de beneficiários exportada com sucesso para:\n{save_path}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o arquivo Excel: {e}")

def exportar_parametros_excel():
    """Exporta a lista de parâmetros de pagamento para um arquivo Excel."""
    save_path = _get_save_path("Relatorio_Parametros_Pagamento.xlsx")
    if not save_path:
        return

    try:
        conn = sqlite3.connect(resource_path('banco.db'))
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                p.id, b.nome_completo, p.cpf_beneficiario, p.valor, p.percentual_concedido,
                p.data_inicial, p.data_final, p.um_terco_ferias, p.salario_13, p.observacoes
            FROM pagamentos p
            JOIN beneficiarios b ON p.beneficiario_id = b.id
            ORDER BY b.nome_completo, p.data_inicial
        """)
        parametros = cursor.fetchall()
        conn.close()

        if not parametros:
            messagebox.showinfo("Informação", "Nenhum parâmetro de pagamento cadastrado para exportar.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Parâmetros de Pagamento"

        headers = ["ID", "Beneficiário", "CPF", "Valor Informado", "Percentual Concedido", "Data Inicial", "Data Final", "1/3 de férias", "13º Salário", "Observações"]
        ws.append(headers)
        _apply_header_style(ws)

        for parametro in parametros:
            parametro_lista = list(parametro)
            parametro_lista[7] = "Sim" if parametro_lista[7] else "Não"
            parametro_lista[8] = "Sim" if parametro_lista[8] else "Não"
            ws.append(parametro_lista)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(save_path)
        messagebox.showinfo("Sucesso", f"Lista de parâmetros exportada com sucesso para:\n{save_path}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o arquivo Excel: {e}")

def exportar_representantes_excel():
    """Exporta a lista completa de representantes legais para um arquivo Excel."""
    save_path = _get_save_path("Relatorio_Representantes_Legais.xlsx")
    if not save_path:
        return

    try:
        conn = sqlite3.connect(resource_path('banco.db'))
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                r.id, r.nome_completo, r.identidade, r.cpf, r.email, r.orgao_emissor, r.endereco, r.telefone,
                r.agencia, r.numero_conta, r.digitoconta, r.tipo_conta, r.numero_banco, b.nome_completo
            FROM representantes_legais r
            LEFT JOIN beneficiarios b ON r.cpf_beneficiario = b.cpf
            ORDER BY r.nome_completo
        """)
        representantes = cursor.fetchall()
        conn.close()

        if not representantes:
            messagebox.showinfo("Informação", "Nenhum representante legal cadastrado para exportar.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Representantes Legais"

        headers = [
            "ID", "Nome do Representante", "Identidade", "CPF", "Email", "Órgão Emissor", "Endereço", "Telefone",
            "Agência", "Conta", "Dígito", "Tipo Conta", "Nº Banco", "Beneficiário Vinculado"
        ]
        ws.append(headers)
        _apply_header_style(ws)

        for representante in representantes:
            ws.append(list(representante))

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(save_path)
        messagebox.showinfo("Sucesso", f"Lista de representantes legais exportada com sucesso para:\n{save_path}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o arquivo Excel: {e}")

def exportar_pagamentos_gerados_excel():
    """Exporta a lista de pagamentos gerados para um arquivo Excel."""
    save_path = _get_save_path("Relatorio_Pagamentos_Gerados.xlsx")
    if not save_path:
        return

    try:
        conn = sqlite3.connect(resource_path('banco.db'))
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                p.id, p.nome_beneficiario, b.cpf, p.mes_referencia, 
                (p.valor - ifnull(p.valor_13_salario, 0) - ifnull(p.valor_um_terco_ferias, 0)), -- Valor Base
                p.valor_13_salario, p.valor_um_terco_ferias, p.valor,
                p.percentual_concedido, p.valor_indice, p.data_geracao, p.data_de_pagamento, b.numero_processo_judicial
            FROM pagamento_gerados p
            LEFT JOIN beneficiarios b ON p.beneficiario_id = b.id
            ORDER BY p.mes_referencia, p.nome_beneficiario
        """)
        pagamentos = cursor.fetchall()
        conn.close()

        if not pagamentos:
            messagebox.showinfo("Informação", "Nenhum pagamento gerado encontrado para exportar.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pagamentos Gerados"

        headers = ["ID", "Beneficiário", "CPF", "Mês Ref.", "Valor Base", "Valor 13º", "Valor Férias", "Valor Total", "Percentual", "Valor Índice", "Data Geração", "Data Pagamento", "Proc. Judicial"]
        ws.append(headers)
        _apply_header_style(ws)

        for pagamento in pagamentos:
            ws.append(pagamento)
        
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if col_idx in [5, 6, 7, 8, 10]:  
                    cell.number_format = '#,##0.00'
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(save_path)
        messagebox.showinfo("Sucesso", f"Lista de pagamentos gerados exportada com sucesso para:\n{save_path}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o arquivo Excel: {e}")